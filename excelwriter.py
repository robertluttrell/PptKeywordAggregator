from openpyxl import Workbook
from openpyxl.styles import Font


class ExcelWriter:

    def __init__(self, excel_path, word_dict):
        self.excel_path = excel_path
        self.word_dict = word_dict
        self.column_map = {"Keyword": 'A', "Filename": 'B', "Count": 'C', "Slide Indices": 'D', "Filepath": 'E'}
        self.workbook = Workbook()
        self.ws = self.workbook.active
        self.label_row_index = 1
        self.last_filled_row_index = None
        self.label_columns()

    def label_columns(self):
        for label in self.column_map.keys():
            cell = self.ws[self.column_map[label] + str(self.label_row_index)]
            cell.font = Font(bold=True)
            cell.value = label

        self.last_filled_row_index = self.label_row_index

    @staticmethod
    def list_to_string(input_list):
        string_list = [str(element) for element in input_list]
        return ','.join(string_list)

    def add_presence_to_sheet(self, word, kfp):
        cur_row = self.last_filled_row_index + 1
        index_list_string = self.list_to_string(kfp.index_list)

        self.ws[self.column_map["Keyword"] + str(cur_row)] = word
        self.ws[self.column_map["Filename"] + str(cur_row)] = kfp.file_name
        self.ws[self.column_map["Count"] + str(cur_row)] = kfp.num_instances
        self.ws[self.column_map["Filepath"] + str(cur_row)] = kfp.file_path
        self.ws[self.column_map["Slide Indices"] + str(cur_row)] = index_list_string

        self.last_filled_row_index = cur_row

    def add_word_to_sheet(self, word, ws):
        """
        Writes all the occurrences of word to the excel worksheet
        :param word: String representing the word to add
        :param ws: Worksheet object representing the worksheet to add to
        """
        for kfp in self.word_dict[word]:
            self.add_presence_to_sheet(word, kfp)

    def write_excel(self):
        """
        Writes the data in df to the excel sheet at self.excel_path
        """
        ws = self.workbook.active
        for word in sorted(self.word_dict.keys()):
            self.add_word_to_sheet(word, ws)

        self.workbook.save(self.excel_path)
